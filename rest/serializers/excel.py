# Standard Library imports
from __future__ import print_function
from datetime import datetime
from tempfile import NamedTemporaryFile

# Package imports
from openpyxl import Workbook
from openpyxl.cell import Cell

# Local imports


def serialize(results):
    uniqueheaders = _get_unique_headers(results)
    workbook = _build_workbook(results, uniqueheaders)
    temp_file = NamedTemporaryFile()
    workbook.save(temp_file)
    temp_file.seek(0)

    return temp_file


def _get_unique_headers(results):
    allheaders = []
    for _, record in enumerate(results):
        allheaders += [key for __, key in enumerate(record)]

    return sorted(set(allheaders))


def _build_workbook(results, headers):
    workbook = Workbook()
    sheet = workbook.worksheets[0]
    sheet.title = 'Sheet1'
    _write_results_to_sheet(headers, results, sheet)
    return workbook


def _write_results_to_sheet(headers, results, sheet, start_row=1):
    for row, record in enumerate(results, start=start_row):
        _validate_headers(record, headers)

        for col, key in enumerate(record, start=1):
            if row == 1:
                _write_col_header(row, col, key, sheet)
            _write_cell_data(row + 1, col, record[key], sheet)


def _write_col_header(row, col, value, sheet):
    cell = sheet.cell(None, row, col)
    cell.style.font.name = 'Arial'
    cell.style.font.size = 10
    cell.style.font.bold = True
    cell.value = value


def _write_cell_data(row, col, value, sheet):
    cell = sheet.cell(None, row, col)
    cell.style.font.name = 'Arial'
    cell.style.font.size = 10

    if isinstance(value, datetime):
        _set_datetime_cell(cell, value)
    else:
        _set_string_cell(cell, value)


def _set_datetime_cell(cell, value):
    cell.value = value
    cell.style.number_format.format_code = '[$-409]m/d/yy\\ h:mm\\ am/pm;@'


def _set_string_cell(cell, value):
    value = _create_excel_string(value)

    cell.set_explicit_value(value, Cell.TYPE_STRING)
    cell.style.number_format = '@'


def _create_excel_string(value):
    if not value:
        return ''
    if isinstance(value, basestring):
        value = value.decode("utf8")
    return ''.join(char for char in unicode(value))


def _validate_headers(uniqueheaders, record):
    for header in uniqueheaders:
        if header not in record:
            raise LookupError('Cannot generate output; missing a header from a record: ' + header)
